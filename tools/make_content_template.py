#!/usr/bin/env python3
"""Buduje wzorzec Excela z tekstami strony (index.html) do wypelnienia przez klienta.

Uzycie:  python3 tools/make_content_template.py [--check]
Wynik:   teksty-strony.xlsx  (arkusze: Instrukcja, Teksty)

ponytail: tekst wyciagany prosto z index.html przez html.parser - zero zaleznosci
poza openpyxl; jesli strona urosnie o kolejne podstrony, dodaj petle po plikach.
"""
import sys
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "index.html"
OUT = ROOT / "teksty-strony.xlsx"

SKIP_TAGS = {"svg", "script", "style"}
VOID = {"br", "img", "meta", "link", "input", "hr", "source"}

SECTIONS = {
    "top": "Hero (pierwszy ekran)",
    "omnie": "O mnie",
    "oferta": "Oferta",
    "narzedzia": "Narzędzia (MTQ / PRISM)",
    "biznes": "Dla biznesu",
    "kontakt": "Kontakt",
    "journey": "Moja droga / Twoja droga",
    "benefits": "Korzyści ze współpracy",
    "site-header": "Menu (góra strony)",
    "site-footer": "Stopka",
}

ELEMENTS = {
    "h1": "Nagłówek główny (H1)",
    "h2": "Nagłówek sekcji (H2)",
    "h3": "Nagłówek karty (H3)",
    "blockquote": "Cytat",
    "li": "Punkt listy",
    "strong": "Wyróżnienie / nazwa",
    "small": "Podpis pod nazwą",
    "span": "Krótki podpis",
    "p": "Akapit",
    "a": "Link",
}


def label(tag, cls):
    if tag == "a":
        return "Przycisk" if "button" in cls else "Link tekstowy"
    if tag == "span" and "eyebrow" in cls:
        return "Nadtytuł (małe litery nad nagłówkiem)"
    if tag == "p" and "hero-lead" in cls:
        return "Zdanie wprowadzające"
    if tag == "p" and "competence-line" in cls:
        return "Linia kompetencji (oddzielona |)"
    if tag == "p" and "tool-kicker" in cls:
        return "Opis narzędzia"
    return ELEMENTS.get(tag, tag)


class Extract(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.stack = []          # [(tag, attrs)]
        self.skip = 0            # glebokosc wewnatrz svg/script/aria-hidden
        self.rows = []           # (sekcja, element, tekst, kontekst)

    # --- pomocnicze -------------------------------------------------
    def section(self):
        for tag, a in reversed(self.stack):
            key = a.get("id") or (a.get("class", "").split() or [""])[0]
            if key in SECTIONS:
                return SECTIONS[key]
        return "SEO i ustawienia ogólne"

    def context(self):
        parts = []
        for tag, a in self.stack[-3:]:
            cls = a.get("class", "").split()
            parts.append(tag + ("." + cls[0] if cls else ""))
        return " > ".join(parts)

    def add(self, element, text, context=None):
        text = " ".join(text.replace("\xa0", " ").split())
        if text:
            self.rows.append((self.section(), element, text, context or self.context()))

    # --- parser -----------------------------------------------------
    def handle_starttag(self, tag, attrs):
        a = dict(attrs)
        if self.skip or tag in SKIP_TAGS or a.get("aria-hidden") == "true":
            self.skip += 1
        elif tag == "img" and a.get("alt"):
            self.add("Opis obrazka (alt — dla wyszukiwarek i czytników ekranu)", a["alt"], a.get("src", ""))
        elif tag == "a" and a.get("href", "").startswith(("http", "mailto:", "tel:")):
            self.add("Adres linku (dokąd prowadzi)", a["href"], self.context())
        elif tag == "meta" and a.get("name") == "description":
            self.add("Opis strony w Google (max ~155 znaków)", a["content"], "meta description")
        if tag not in VOID:
            self.stack.append((tag, a))

    def handle_startendtag(self, tag, attrs):   # np. <path/> w <svg>
        self.handle_starttag(tag, attrs)
        self.handle_endtag(tag)

    def handle_endtag(self, tag):
        if self.skip:
            self.skip -= 1
        if self.stack and tag not in VOID:
            self.stack.pop()

    def handle_data(self, data):
        if self.skip or not self.stack or not data.strip():
            return
        tag, a = self.stack[-1]
        if tag in ("head", "body", "html"):
            return
        cls = a.get("class", "")
        if tag == "title":
            self.add("Tytuł strony w karcie przeglądarki i Google", data, "<title>")
        else:
            self.add(label(tag, cls), data)


def build(rows):
    wb = Workbook()

    info = wb.active
    info.title = "Instrukcja"
    info.column_dimensions["A"].width = 110
    for i, line in enumerate([
        "WZORZEC TEKSTÓW — strona Katarzyna Chałas",
        "",
        "1. Wypełnij TYLKO kolumnę \"NOWY TEKST\" w arkuszu \"Teksty\".",
        "2. Puste pole = zostaje tekst obecny. Wpis \"USUŃ\" = element znika ze strony.",
        "3. Nie zmieniaj kolumny ID — po niej podstawiamy teksty na stronie.",
        "4. Kolumna \"Limit znaków\" to sugestia: dłuższy tekst też się zmieści, ale może rozjechać układ.",
        "5. Jeden wiersz = jeden element na stronie (nagłówek, akapit, przycisk, punkt listy).",
        "6. Wiersze \"Opis obrazka (alt)\" widzą tylko wyszukiwarki i czytniki ekranu — krótko, rzeczowo.",
        "7. Wiersze \"Adres linku\" to docelowe adresy (np. LinkedIn, e-mail: mailto:adres@domena.pl).",
        "8. Uwagi, wątpliwości i prośby o nowe sekcje wpisuj w kolumnę \"UWAGI\".",
        "9. Odeślij plik w formacie .xlsx — resztę robimy po naszej stronie.",
    ], start=1):
        c = info.cell(row=i, column=1, value=line)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if i == 1:
            c.font = Font(bold=True, size=14)

    ws = wb.create_sheet("Teksty")
    headers = ["ID", "Sekcja strony", "Rodzaj elementu", "OBECNY TEKST",
               "Limit znaków", "NOWY TEKST (wypełnij)", "UWAGI", "Element HTML"]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="2F4858")
    fill_new = PatternFill("solid", fgColor="FFF6E5")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 30

    for i, (section, element, text, ctx) in enumerate(rows, start=1):
        limit = max(40, int(len(text) * 1.25) // 5 * 5)
        ws.append([f"T{i:03d}", section, element, text, limit, "", "", ctx])

    widths = [8, 24, 30, 60, 11, 60, 26, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = border
            c.alignment = Alignment(wrap_text=True, vertical="top")
        row[5].fill = fill_new
        row[6].fill = fill_new
        row[7].font = Font(size=8, color="909090")

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    wb.save(OUT)
    return ws.max_row - 1


def check(rows):
    texts = [r[2] for r in rows]
    assert "Szkolenia, mentoring i coaching dla liderek, liderów i zespołów" in texts, "brak H1"
    assert "MTQ" not in texts  # <text> w ikonie SVG, "tekst z <svg> nie powinien trafic do wzorca"
    assert "→" not in texts, "elementy aria-hidden pomijamy"
    assert any(t.startswith("Katarzyna Chałas — psycholożka") for t in texts), "brak meta description"
    assert sum(1 for r in rows if r[0] == "Oferta") >= 19, "sekcja Oferta niekompletna"
    assert len(rows) > 60, f"za malo tekstow: {len(rows)}"
    print(f"OK — {len(rows)} tekstow, sekcje: {sorted({r[0] for r in rows})}")


if __name__ == "__main__":
    p = Extract()
    p.feed(SRC.read_text(encoding="utf-8"))
    if "--check" in sys.argv:
        check(p.rows)
    else:
        print(f"{OUT.name}: {build(p.rows)} wierszy")
